from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException
from datetime import datetime, timedelta
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
from . import models
from .database import SessionLocal
import os
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from email.mime.application import MIMEApplication
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback
import socket
from sqlalchemy.orm import Session
import ssl  # SSL 컨텍스트 사용을 위한 모듈 추가

# 프로젝트 루트 디렉토리 찾기
BASE_DIR = Path(__file__).resolve().parent.parent

# .env 파일 명시적 경로 지정
load_dotenv(os.path.join(BASE_DIR, ".env"))

# 이메일 발송 설정
EMAIL_ID = os.getenv('SMTP_USER')
EMAIL_PW = os.getenv('SMTP_PASSWORD')
EMAIL_SENDER = os.getenv('SMTP_USER', 'ckdpharmamorning@gmail.com')

# 파일 상단에 글로벌 설정 변수 추가
# False로 설정하면 디버그 로그가 표시되지 않음
# 로그가 필요할 때만 True로 변경
DEBUG_SMTP = False  # 여기서 한 번만 설정하면 됨

# 웹드라이버 설정 및 대기 시간 설정
DEFAULT_TIMEOUT = 10  # 기본 대기 시간(초)
PAGE_LOAD_TIMEOUT = 20  # 페이지 로딩 타임아웃(초)

def setup_driver():
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, DEFAULT_TIMEOUT)
    driver.implicitly_wait(5)  # 암시적 대기 시간 증가
    driver.maximize_window()
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)  # 페이지 로드 타임아웃 설정
    return driver, wait

def handle_popups(driver):
    """일반적인 팝업을 처리하는 함수"""
    try:
        # 팝업 닫기 버튼들의 일반적인 속성들
        popup_close_patterns = [
            (By.XPATH, "//button[contains(text(), '닫기')]"),
            (By.XPATH, "//a[contains(text(), '닫기')]"),
            (By.XPATH, "//button[contains(text(), '취소')]"),
            (By.XPATH, "//button[contains(text(), 'Close')]"),
            (By.XPATH, "//button[@class='close']"),
            (By.XPATH, "//button[contains(@class, 'popup-close')]"),
            (By.XPATH, "//div[contains(@class, 'popup')]//button"),
            (By.XPATH, "//div[contains(@class, 'modal')]//button"),
            (By.XPATH, "//span[contains(text(), '×')]"),
            (By.XPATH, "//button[contains(@class, 'btn-close')]"),
            (By.CSS_SELECTOR, ".popup .close"),
            (By.CSS_SELECTOR, ".modal .close"),
            (By.CSS_SELECTOR, ".popup-container .close"),
            (By.CSS_SELECTOR, ".modal-content .close")
        ]

        for selector_type, selector in popup_close_patterns:
            try:
                # 짧은 대기 시간으로 요소 찾기 시도
                close_buttons = driver.find_elements(selector_type, selector)
                for button in close_buttons:
                    if button.is_displayed():
                        print(f"팝업 닫기 버튼 발견: {selector}")
                        button.click()
                        time.sleep(0.5)  # 팝업이 닫히는 데 시간이 필요할 수 있음
            except (NoSuchElementException, ElementNotInteractableException):
                continue

        # 알림 창 수락
        try:
            alert = driver.switch_to.alert
            alert.accept()
            print("알림 창을 닫았습니다.")
        except:
            pass  # 알림 창이 없는 경우 무시

        return True
    except Exception as e:
        print(f"팝업 처리 중 오류 발생: {str(e)}")
        return False

def wait_for_page_load(driver, url, timeout=PAGE_LOAD_TIMEOUT):
    """페이지가 완전히 로드될 때까지 대기하는 함수"""
    try:
        print(f"{url} 페이지 로딩 중...")
        driver.get(url)
        
        # 페이지 로드 완료 대기
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        # 잠시 대기하여 동적 콘텐츠가 로드될 시간 제공
        time.sleep(1)
        
        # 가능한 팝업 처리
        handle_popups(driver)
        
        print(f"{url} 페이지 로딩 완료")
        return True
    except TimeoutException:
        print(f"{url} 페이지 로드 시간 초과")
        return False
    except Exception as e:
        print(f"{url} 페이지 로드 중 오류 발생: {str(e)}")
        return False

# 로깅 유틸리티 함수 추가
def log_message(message, log_type="INFO"):
    """
    타임스탬프와 로그 유형을 포함한 로그 메시지를 출력하는 유틸리티 함수
    
    Args:
        message: 로그 메시지
        log_type: 로그 유형 (INFO, WARNING, ERROR, SUCCESS 등)
    """
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    
    # 로그 유형에 따른 색상 코드 (터미널에서만 작동)
    colors = {
        "INFO": "",      # 기본색
        "WARNING": "\033[93m",  # 노란색
        "ERROR": "\033[91m",    # 빨간색
        "SUCCESS": "\033[92m",  # 초록색
        "DEBUG": "\033[94m"     # 파란색
    }
    
    reset_color = "\033[0m"  # 색상 리셋
    
    # 색상 적용 (지원되는 경우)
    color_code = colors.get(log_type, "")
    reset = reset_color if color_code else ""
    
    try:
        print(f"[{timestamp}] {color_code}{log_type}: {message}{reset}")
    except:
        # 색상 코드가 지원되지 않는 환경에서는 일반 텍스트로 출력
        print(f"[{timestamp}] {log_type}: {message}")

def capture_screenshot(driver, name="error"):
    """
    현재 브라우저 상태의 스크린샷을 캡처하는 함수
    
    Args:
        driver: Selenium 웹드라이버 인스턴스
        name: 스크린샷 파일명 접두사
        
    Returns:
        저장된 스크린샷 파일 경로 또는 None (실패 시)
    """
    try:
        # 스크린샷 저장 디렉토리 생성
        screenshot_dir = os.path.join(BASE_DIR, "screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        
        # 타임스탬프를 포함한 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{name}_{timestamp}.png"
        filepath = os.path.join(screenshot_dir, filename)
        
        # 스크린샷 저장
        driver.save_screenshot(filepath)
        log_message(f"스크린샷 저장 완료: {filepath}", "INFO")
        
        return filepath
    except Exception as e:
        log_message(f"스크린샷 캡처 실패: {str(e)}", "ERROR")
        return None

def collect_dailypharm_headlines(driver, wait):
    log_message("데일리팜 헤드라인 수집 시작", "INFO")
    headlines = []
    start_time = datetime.now()
    
    try:
        # 메인 페이지 접속 - 향상된 대기 로직 사용
        url = "https://www.dailypharm.com/"
        log_message(f"데일리팜 URL 접속 시도: {url}", "INFO")
        
        if not wait_for_page_load(driver, url):
            log_message("데일리팜 페이지 로드 실패, 헤드라인 수집 건너뜀", "ERROR")
            capture_screenshot(driver, "dailypharm_load_failed")
            return []
        
        log_message("데일리팜 페이지 로드 성공", "SUCCESS")
        
        # 지정된 CSS 선택자로 헤드라인 가져오기
        selectors = [
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div:nth-child(1) > div > div.bn_left > div.main_banner_main_t > em > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(1) > div:nth-child(2) > div.dpfocus_title > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(3) > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(4) > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(5) > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(6) > a",
            "body > div.mainView > div:nth-child(2) > div.MainHeadLine > div.center > div.TopHead.d_TopHead_2 > div.subTop > ul > li:nth-child(7) > a"
        ]
        
        log_message(f"데일리팜 - {len(selectors)}개 선택자에서 헤드라인 수집 시도", "INFO")
        
        for idx, selector in enumerate(selectors, 1):
            try:
                # 명시적 대기로 요소 찾기 시도
                log_message(f"데일리팜 - 선택자 {idx}/{len(selectors)} 처리 중: {selector[:30]}...", "DEBUG")
                try:
                    element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                except TimeoutException:
                    log_message(f"선택자 '{selector[:30]}...' 요소를 찾을 수 없습니다 (타임아웃).", "WARNING")
                    continue
                
                headline = element.text.strip()
                url = element.get_attribute('href') if element.tag_name == 'a' else ""
                
                if headline:
                    headlines.append({
                        'source': 'dailypharm',
                        'headline': headline,
                        'url': url,
                        'published_at': datetime.now()
                    })
                    log_message(f"데일리팜 헤드라인 수집: {headline}", "SUCCESS")
                else:
                    log_message(f"데일리팜 - 선택자 {idx} 요소 발견했으나 헤드라인 텍스트가 비어 있음", "WARNING")
            except Exception as e:
                log_message(f"선택자 '{selector[:30]}...' 처리 중 오류: {str(e)}", "ERROR")
        
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        log_message(f"데일리팜 헤드라인 총 {len(headlines)}건 수집 완료 ({duration:.2f}초 소요)", "SUCCESS")
        
        if len(headlines) == 0:
            log_message("데일리팜에서 헤드라인을 찾지 못했습니다. 페이지 구조가 변경되었을 수 있습니다.", "WARNING")
            capture_screenshot(driver, "dailypharm_no_headlines")
        
        return headlines
            
    except Exception as e:
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        log_message(f"데일리팜 헤드라인 수집 중 에러 발생: {str(e)} ({duration:.2f}초 소요)", "ERROR")
        log_message(f"스택 트레이스: {traceback.format_exc()}", "ERROR")
        capture_screenshot(driver, "dailypharm_error")
        return []

def collect_yakup_headlines(driver, wait):
    print("\n=== 약업닷컴 헤드라인 수집 시작 ===")
    headlines = []
    
    try:
        # 메인 페이지 접속 - 향상된 대기 로직 사용
        url = "https://yakup.com/"
        if not wait_for_page_load(driver, url):
            print("약업닷컴 페이지 로드 실패, 헤드라인 수집 건너뜀")
            return []
        
        # 지정된 CSS 선택자로 헤드라인 가져오기
        selectors = [
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.main_banner_con.tmp > a > div.text_con > div.title_con > span",
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.main_article_con > a:nth-child(1) > table > tbody > tr > td.text_td > div > div.title_con > span",
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.main_article_con > a:nth-child(2) > table > tbody > tr > td.text_td > div > div.title_con > span",
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.main_article_con > a:nth-child(3) > table > tbody > tr > td.text_td > div > div.title_con > span",
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.list_con.cf > div.left_con > div > div.left_list02_con > ul.text_1_ul > li:nth-child(1) > a > div > table > tbody > tr > td.text_td > div.title_con > span",
            "#main_con > div.contents_con.cf > div.left_con.layout_left_con > div > div.list_con.cf > div.left_con > div > div.left_list02_con > ul.text_1_ul > li:nth-child(2) > a > div > table > tbody > tr > td.text_td > div.title_con > span"
        ]
        
        for selector in selectors:
            try:
                # 명시적 대기로 요소 찾기 시도
                try:
                    element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                except TimeoutException:
                    print(f"선택자 '{selector}' 요소를 찾을 수 없습니다 (타임아웃).")
                    continue
                
                headline = element.text.strip()
                
                # 상위 a 태그에서 URL 가져오기
                parent_element = element
                while parent_element and parent_element.tag_name != 'a':
                    parent_element = parent_element.find_element(By.XPATH, '..')
                
                url = parent_element.get_attribute('href') if parent_element and parent_element.tag_name == 'a' else ""
                
                if headline:
                    headlines.append({
                        'source': 'yakup',
                        'headline': headline,
                        'url': url,
                        'published_at': datetime.now()
                    })
                    print(f"약업닷컴 헤드라인 수집: {headline}")
            except Exception as e:
                print(f"선택자 '{selector}' 처리 중 오류: {str(e)}")
        
        print(f"약업닷컴 헤드라인 총 {len(headlines)}건 수집 완료")
        return headlines
            
    except Exception as e:
        print(f"약업닷컴 헤드라인 수집 중 에러 발생: {str(e)}")
        print(f"스택 트레이스: {traceback.format_exc()}")
        return []

def collect_hitnews_headlines(driver, wait):
    print("\n=== 히트뉴스 헤드라인 수집 시작 ===")
    headlines = []
    
    try:
        # 메인 페이지 접속 - 향상된 대기 로직 사용
        url = "http://www.hitnews.co.kr/"
        if not wait_for_page_load(driver, url):
            print("히트뉴스 페이지 로드 실패, 헤드라인 수집 건너뜀")
            return []
        
        # 지정된 CSS 선택자로 헤드라인 가져오기
        selectors = [
            "#skin-67 > div.item.large > a > span.content.for-middle > strong",
            "#skin-67 > div:nth-child(2) > a > span.content > strong",
            "#skin-67 > div:nth-child(3) > a > span.content > strong",
            "#skin-67 > div:nth-child(4) > a > span.content > strong",
            "#skin-13 > div:nth-child(1) > a",
            "#skin-13 > div:nth-child(2) > a"
        ]
        
        for selector in selectors:
            try:
                # 명시적 대기로 요소 찾기 시도
                try:
                    element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                except TimeoutException:
                    print(f"선택자 '{selector}' 요소를 찾을 수 없습니다 (타임아웃).")
                    continue
                
                headline = element.text.strip()
                
                # a 태그인 경우 직접 URL 가져오기, 아닌 경우 상위 a 태그 찾기
                if element.tag_name == 'a':
                    url = element.get_attribute('href')
                else:
                    parent_element = element
                    while parent_element and parent_element.tag_name != 'a':
                        parent_element = parent_element.find_element(By.XPATH, '..')
                    url = parent_element.get_attribute('href') if parent_element and parent_element.tag_name == 'a' else ""
                
                if headline:
                    headlines.append({
                        'source': 'hitnews',
                        'headline': headline,
                        'url': url,
                        'published_at': datetime.now()
                    })
                    print(f"히트뉴스 헤드라인 수집: {headline}")
            except Exception as e:
                print(f"선택자 '{selector}' 처리 중 오류: {str(e)}")
        
        print(f"히트뉴스 헤드라인 총 {len(headlines)}건 수집 완료")
        return headlines
            
    except Exception as e:
        print(f"히트뉴스 헤드라인 수집 중 에러 발생: {str(e)}")
        print(f"스택 트레이스: {traceback.format_exc()}")
        return []

def collect_kpanews_headlines(driver, wait):
    print("\n=== 약사공론 헤드라인 수집 시작 ===")
    headlines = []
    
    try:
        # 메인 페이지 접속 - 향상된 대기 로직 사용
        url = "https://www.kpanews.co.kr/"
        if not wait_for_page_load(driver, url):
            print("약사공론 페이지 로드 실패, 헤드라인 수집 건너뜀")
            return []
        
        # 지정된 CSS 선택자로 헤드라인 가져오기
        selectors = [
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.msec.msec1 > div > section.headline_news.headline_news_wcm > div > div > p > a.t-dot-multi",
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.msec.msec1 > div > section.headline_news_list.headline_news_list_wcc.wctm_mainre1 > dl > dt:nth-child(1) > ul > li:nth-child(1) > a > div.tit > p",
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.msec.msec1 > div > section.headline_news_list.headline_news_list_wcc.wctm_mainre1 > dl > dt:nth-child(1) > ul > li:nth-child(2) > a > div.tit > p",
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.mwide > section > dl > dt:nth-child(1) > ul > li:nth-child(1) > div > p.t1.t-dot-multi > a",
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.mwide > section > dl > dt:nth-child(1) > ul > li:nth-child(2) > div > p.t1.t-dot-multi > a",
            "#container > div.inr-c > div:nth-child(5) > div.lft_m > div.mwide > section > dl > dt:nth-child(1) > ul > li:nth-child(3) > div > p.t1.t-dot-multi > a"
        ]
        
        for selector in selectors:
            try:
                # 명시적 대기로 요소 찾기 시도
                try:
                    element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                except TimeoutException:
                    print(f"선택자 '{selector}' 요소를 찾을 수 없습니다 (타임아웃).")
                    continue
                
                headline = element.text.strip()
                
                # a 태그인 경우 직접 URL 가져오기, 아닌 경우 상위 a 태그 찾기
                if element.tag_name == 'a':
                    url = element.get_attribute('href')
                else:
                    parent_element = element
                    while parent_element and parent_element.tag_name != 'a':
                        parent_element = parent_element.find_element(By.XPATH, '..')
                    url = parent_element.get_attribute('href') if parent_element and parent_element.tag_name == 'a' else ""
                
                if headline:
                    headlines.append({
                        'source': 'kpanews',
                        'headline': headline,
                        'url': url,
                        'published_at': datetime.now()
                    })
                    print(f"약사공론 헤드라인 수집: {headline}")
            except Exception as e:
                print(f"선택자 '{selector}' 처리 중 오류: {str(e)}")
        
        print(f"약사공론 헤드라인 총 {len(headlines)}건 수집 완료")
        return headlines
            
    except Exception as e:
        print(f"약사공론 헤드라인 수집 중 에러 발생: {str(e)}")
        print(f"스택 트레이스: {traceback.format_exc()}")
        return []

# 4개 언론사의 헤드라인을 모두 수집하는 통합 함수
def collect_all_headlines(driver, wait, max_retries=3, retry_delay=5):
    """
    4개 언론사의 헤드라인을 수집하는 통합 함수
    
    Args:
        driver: Selenium 웹드라이버 인스턴스
        wait: WebDriverWait 인스턴스
        max_retries: 수집 실패 시 최대 재시도 횟수 (기본값: 3)
        retry_delay: 재시도 간 대기 시간(초) (기본값: 5)
        
    Returns:
        수집된 모든 헤드라인 목록
    """
    print("\n" + "="*50)
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 4개 언론사 헤드라인 수집 시작")
    print("="*50)
    
    # 각 사이트별 수집 결과와 시도 횟수를 추적하는 로깅 데이터
    collection_stats = {
        'dailypharm': {'success': False, 'attempts': 0, 'count': 0, 'start_time': None, 'end_time': None},
        'yakup': {'success': False, 'attempts': 0, 'count': 0, 'start_time': None, 'end_time': None},
        'hitnews': {'success': False, 'attempts': 0, 'count': 0, 'start_time': None, 'end_time': None},
        'kpanews': {'success': False, 'attempts': 0, 'count': 0, 'start_time': None, 'end_time': None}
    }
    
    # 최종 수집된 헤드라인 목록
    all_headlines = []
    
    # 각 사이트별 수집 함수와 사이트명을 매핑
    sites = [
        {'name': 'dailypharm', 'func': collect_dailypharm_headlines, 'korean_name': '데일리팜'},
        {'name': 'yakup', 'func': collect_yakup_headlines, 'korean_name': '약업닷컴'},
        {'name': 'hitnews', 'func': collect_hitnews_headlines, 'korean_name': '히트뉴스'},
        {'name': 'kpanews', 'func': collect_kpanews_headlines, 'korean_name': '약사공론'}
    ]
    
    # 각 사이트에서 헤드라인 수집 (재시도 로직 포함)
    for site in sites:
        site_name = site['name']
        site_korean_name = site['korean_name']
        collection_func = site['func']
        
        print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 시작")
        collection_stats[site_name]['start_time'] = datetime.now()
        
        headlines = []
        attempts = 0
        success = False
        
        # 최대 재시도 횟수만큼 시도
        while attempts < max_retries and not success:
            attempts += 1
            collection_stats[site_name]['attempts'] = attempts
            
            try:
                # 시도 횟수가 1보다 크면 재시도 메시지 출력
                if attempts > 1:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} {attempts}번째 시도 중...")
                
                # 헤드라인 수집 시도
                headlines = collection_func(driver, wait)
                
                # 결과 확인 및 로깅
                if headlines and len(headlines) > 0:
                    success = True
                    collection_stats[site_name]['success'] = True
                    collection_stats[site_name]['count'] = len(headlines)
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 성공: {len(headlines)}건")
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 실패: 헤드라인이 없습니다.")
                    
                    # 마지막 시도가 아니면 잠시 대기 후 재시도
                    if attempts < max_retries:
                        # 지수 백오프: 시도 횟수에 따라 대기 시간 증가
                        wait_time = retry_delay * (2 ** (attempts - 1))
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {wait_time}초 후 재시도합니다.")
                        time.sleep(wait_time)
                        
                        # 페이지 새로고침 시도
                        try:
                            driver.refresh()
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 페이지 새로고침 완료")
                        except Exception as refresh_err:
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 페이지 새로고침 실패: {str(refresh_err)}")
                    else:
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 최대 시도 횟수 도달. 수집 포기.")
            
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 중 오류 발생: {str(e)}")
                print(f"Stack trace: {traceback.format_exc()}")
                
                # 마지막 시도가 아니면 잠시 대기 후 재시도
                if attempts < max_retries:
                    wait_time = retry_delay * (2 ** (attempts - 1))
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {wait_time}초 후 재시도합니다.")
                    time.sleep(wait_time)
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 최대 시도 횟수 도달. 수집 포기.")
        
        # 수집 종료 시간 기록
        collection_stats[site_name]['end_time'] = datetime.now()
        duration = (collection_stats[site_name]['end_time'] - collection_stats[site_name]['start_time']).total_seconds()
        
        # 결과에 따른 로그 메시지
        if success:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 완료: {len(headlines)}건 ({duration:.1f}초 소요, {attempts}번 시도)")
            # 헤드라인 추가
            all_headlines.extend(headlines)
        else:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {site_korean_name} 수집 실패: 헤드라인을 가져오지 못했습니다. ({duration:.1f}초 소요, {attempts}번 시도)")
    
    # 전체 결과 요약 출력
    print("\n" + "="*50)
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 4개 언론사 헤드라인 수집 결과 요약")
    print("-"*50)
    
    total_count = len(all_headlines)
    success_sites = sum(1 for stats in collection_stats.values() if stats['success'])
    total_attempts = sum(stats['attempts'] for stats in collection_stats.values())
    
    for site_name, stats in collection_stats.items():
        site_korean = next((site['korean_name'] for site in sites if site['name'] == site_name), site_name)
        status = "✓ 성공" if stats['success'] else "✗ 실패"
        duration = (stats['end_time'] - stats['start_time']).total_seconds() if stats['end_time'] and stats['start_time'] else 0
        print(f"{site_korean}: {status} | {stats['count']}건 | {stats['attempts']}번 시도 | {duration:.1f}초 소요")
    
    print("-"*50)
    print(f"총 {total_count}건 수집 완료 (성공 사이트: {success_sites}/4, 총 시도 횟수: {total_attempts})")
    print("="*50)
    
    # 수집된 헤드라인이 없으면 전체 재시도
    if total_count == 0:
        print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 모든 사이트에서 헤드라인을 수집하지 못했습니다. 전체 재시도를 시작합니다.")
        # 브라우저 재시작
        try:
            driver.quit()
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 웹드라이버 종료")
            time.sleep(3)
            
            # 새로운 드라이버 생성
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 새 웹드라이버 생성 중...")
            driver, wait = setup_driver()
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 새 웹드라이버 생성 완료")
            
            # 재귀적으로 함수 호출 (단, 재시도 횟수는 1로 제한하여 무한 루프 방지)
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 전체 수집 과정 재시도 중...")
            all_headlines = collect_all_headlines(driver, wait, max_retries=1, retry_delay=retry_delay)
        except Exception as restart_err:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 웹드라이버 재시작 중 오류 발생: {str(restart_err)}")
            print(f"Stack trace: {traceback.format_exc()}")
    
    return all_headlines

# 통계 정보 생성 함수
def generate_stats_content(all_headlines):
    # 오늘과 어제 날짜 계산
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    
    # 하드코딩된 8시로 명확하게 정의
    collection_start = f"{yesterday.strftime('%Y년 %m월 %d일')} 08:00"
    collection_end = f"{today.strftime('%Y년 %m월 %d일')} 08:00"
    
    # 수집 기간 문자열 생성 (고정된 08:00~08:00 사용)
    collection_period = f"{collection_start} ~ {collection_end} 까지의 뉴스"
    
    # 출처별 뉴스 수 계산
    sources = {
        'dailypharm': {'name': '데일리팜', 'count': 0},
        'yakup': {'name': '약업닷컴', 'count': 0},
        'hitnews': {'name': '히트뉴스', 'count': 0},
        'kpanews': {'name': '약사공론', 'count': 0}
    }
    
    for headline in all_headlines:
        if headline['source'] in sources:
            sources[headline['source']]['count'] += 1
    
    # 총 뉴스 건수 계산
    total_count = sum(source['count'] for source in sources.values())
    
    # 소스별 뉴스 수 문자열 생성
    source_counts = ", ".join([f"{source['name']} {source['count']}건" for source in sources.values() if source['count'] > 0])
    
    # 통계 정보 HTML 생성
    stats_content = f"""
    <div style="margin-bottom: 20px; font-family: Arial, sans-serif;">
        <div style="text-align: right; font-size: 12px; color: #666; margin-bottom: 10px;">
            <p>🕒 수집 기간: {collection_period}</p>
            <p>📚 총 {total_count}건 ({source_counts})</p>
        </div>
        <h2 style="color: #4b77be; margin-bottom: 10px;">{today.strftime("%Y년 %m월 %d일")} 헤드라인</h2>
        <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px;">
            <table style="width: 100%; border-collapse: collapse;">
                <tr>
                    <th style="width: 20%; text-align: left; padding: 8px; border-bottom: 1px solid #ddd;">출처</th>
                    <th style="width: 80%; text-align: left; padding: 8px; border-bottom: 1px solid #ddd;">주요 헤드라인</th>
                </tr>
    """
    
    # 각 출처별 뉴스 그룹화
    for source_key, source_info in sources.items():
        if source_info['count'] > 0:
            source_headlines = [h for h in all_headlines if h['source'] == source_key]
            
            stats_content += f"""
                <tr>
                    <td style="vertical-align: top; padding: 8px; border-bottom: 1px solid #ddd; font-weight: bold;">{source_info['name']}</td>
                    <td style="vertical-align: top; padding: 8px; border-bottom: 1px solid #ddd;">
                        <ul style="margin: 0; padding-left: 20px;">
            """
            
            for headline in source_headlines:
                url = headline.get('url', '')
                if url:
                    stats_content += f'<li style="margin-bottom: 10px;"><a href="{url}" style="color: #1a0dab; text-decoration: none;">{headline["headline"]}</a></li>'
                else:
                    stats_content += f'<li style="margin-bottom: 10px;">{headline["headline"]}</li>'
            
            stats_content += """
                        </ul>
                    </td>
                </tr>
            """
    
    stats_content += """
            </table>
        </div>
    """
    
    
    return stats_content

# 이메일 발송 함수
def send_email(subscriber, subject, content, attachment_path=None, attachment_filename=None):
    try:
        # 도메인 설정 (환경변수에서 가져오기)
        domain = os.environ.get('APP_DOMAIN', '10.4.31.151:80')
        
        # 구독 취소 링크 생성
        unsubscribe_url = f"http://{domain}/unsubscribe/{subscriber.unsubscribe_token}"
        
        # 중괄호 형식 변경 및 직접 대체
        content = content.replace("{{ unsubscribe_link }}", unsubscribe_url)
        content = content.replace("{ unsubscribe_link }", unsubscribe_url)  # 여백이 있는 경우도 처리
        
        # 멀티파트 메시지 생성 (mixed 타입으로 멀티파트 메시지 생성)
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = EMAIL_SENDER
        msg['To'] = subscriber.email
        
        # 구독 취소 헤더 추가 (일부 이메일 클라이언트가 지원)
        msg.add_header('List-Unsubscribe', f'<{unsubscribe_url}>')
        
        # HTML 메시지 첨부
        html_part = MIMEText(content, 'html', 'utf-8')
        msg.attach(html_part)
        
        # 첨부 파일이 있으면 추가
        if attachment_path and os.path.exists(attachment_path):
            print(f"첨부 파일 처리 시작: {attachment_path}")
            
            # 파일명이 지정되지 않은 경우 원본 파일명 사용
            if not attachment_filename:
                attachment_filename = os.path.basename(attachment_path)
            
            print(f"첨부 파일명: {attachment_filename}")
            
            # 엑셀 파일 첨부 (MIMEApplication 사용)
            try:
                with open(attachment_path, 'rb') as file:
                    attachment_data = file.read()
                    print(f"파일 크기: {len(attachment_data)} 바이트")
                    
                    # 엑셀 파일은 MIMEApplication으로 첨부
                    attachment = MIMEApplication(attachment_data, _subtype='xlsx')
                    attachment.add_header('Content-Disposition', 'attachment', 
                                        filename=attachment_filename)
                    msg.attach(attachment)
                    print(f"엑셀 파일 첨부 성공: {attachment_filename}")
            except Exception as attach_err:
                print(f"파일 첨부 중 오류 발생: {str(attach_err)}")
                print(traceback.format_exc())
        
        # 이메일 발송 - 소켓 연결 유지 설정 추가
        with smtplib.SMTP('smtp.gmail.com', 587, timeout=30) as server:
            # 전역 변수로 디버그 모드 제어
            if DEBUG_SMTP:
                server.set_debuglevel(1)
            else:
                server.set_debuglevel(0)
            
            # 소켓 연결 유지 설정 (기존 코드 그대로 유지)
            server.sock.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
            
            # TCP keepalive 파라미터 설정 (리눅스) (기존 코드 그대로 유지)
            if hasattr(socket, 'TCP_KEEPIDLE') and hasattr(socket, 'TCP_KEEPINTVL') and hasattr(socket, 'TCP_KEEPCNT'):
                server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPIDLE, 60)
                server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPINTVL, 60)
                server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPCNT, 5)
            
            server.ehlo()
            # 여기만 수정: SSL 컨텍스트 명시적 사용
            context = ssl.create_default_context()
            server.starttls(context=context)
            server.ehlo()
            server.login(EMAIL_ID, EMAIL_PW)
            server.send_message(msg)
        
        print(f"이메일 발송 성공: {subscriber.email}")
        return True
    except Exception as e:
        print(f"이메일 발송 실패 ({subscriber.email}): {str(e)}")
        print(f"스택 트레이스: {traceback.format_exc()}")
        return False

# 이메일 본문 생성 함수 전체 수정
def generate_email_content(all_headlines):
    # 헤드라인을 최신순으로 정렬
    all_headlines = sorted(all_headlines, key=lambda x: x.get('date', ''), reverse=True)
    
    # 오늘과 어제 날짜 계산
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    
    # 수집 기간 문자열 생성 (어제 9시 ~ 오늘 9시)
    collection_period = f"{yesterday.strftime('%Y년 %m월 %d일')} 08:00 ~ {today.strftime('%Y년 %m월 %d일')} 08:00 까지의 뉴스"
    
    # 통계 정보 계산
    total_count = len(all_headlines)
    
    # 출처별 뉴스 수 계산
    source_counts = {}
    for headline in all_headlines:
        source = headline['source']
        if source not in source_counts:
            source_counts[source] = 0
        source_counts[source] += 1
    
    # 통계 정보 헤더 생성 - 테이블과 정확히 같은 위치에 맞춤
    stats_header = f"""
    <div style="width: 80%; max-width: 800px; min-width: 500px; margin: 0 auto 15px; padding: 0; box-sizing: border-box; font-size: 13px; color: #666;">
        <div style="text-align: right; padding-right: 0;">
            수집 기간: {collection_period}<br>
            📰 총 {total_count}건 (
    """
    
    # 출처별 통계 추가
    source_stats = []
    for source, count in source_counts.items():
        # 출처 이름 한글로 변환
        korean_source = source
        if source == "dailypharm":
            korean_source = "데일리팜"
        elif source == "yakup":
            korean_source = "약업신문"
        elif source == "hitnews":
            korean_source = "히트뉴스"
        elif source == "kpanews":
            korean_source = "약사공론"
            
        source_stats.append(f"{korean_source} {count}건")
    
    stats_header += ", ".join(source_stats) + ")</div></div>"
    
    # 테이블 헤더 정렬 수정 - 제목 열 가운데 정렬로 변경
    table_content = """
    <table border="1" cellspacing="0" cellpadding="12" style="width: 80%; max-width: 800px; min-width: 500px; border-collapse: collapse; border: 1px solid #dddddd; font-family: Arial, sans-serif; font-size: 15px; margin: 0 auto;">
        <tr style="background-color: #4472C4; color: white; height: 45px;">
            <th style="width: 8%; text-align: center; border: 1px solid #dddddd; font-weight: normal;">No.</th>
            <th style="width: 62%; text-align: center; border: 1px solid #dddddd; font-weight: normal;">제목</th>
            <th style="width: 10%; text-align: center; border: 1px solid #dddddd; font-weight: normal;">링크</th>
            <th style="width: 20%; text-align: center; border: 1px solid #dddddd; font-weight: normal;">출처</th>
        </tr>
    """
    
    # 중복 제거를 위한 사용된 제목/링크 추적
    used_titles = set()
    
    # 각 헤드라인에 대한 테이블 행 생성
    counter = 1
    for headline in all_headlines:
        title = headline['headline']
        link = headline['url']
        
        # 중복 제목 건너뛰기
        if title in used_titles:
            continue
            
        used_titles.add(title)
        source = headline['source']
        
        # 출처 이름 한글로 변환
        if source == "dailypharm":
            source = "데일리팜"
        elif source == "yakup":
            source = "약업신문"
        elif source == "hitnews":
            source = "히트뉴스"
        elif source == "kpanews":
            source = "약사공론"
        
        # 배경색 설정 (짝수/홀수 행)
        bg_color = "#f9f9f9" if counter % 2 == 0 else "#ffffff"
        
        # 테이블 행 부분 - 제목 열에 왼쪽 정렬 유지하고 패딩 추가
        table_content += f"""
        <tr style="background-color: {bg_color}; height: 40px;">
            <td style="text-align: center; border: 1px solid #dddddd;">{counter}</td>
            <td style="text-align: left; border: 1px solid #dddddd; line-height: 1.4; word-break: break-word; padding-left: 15px;">{title}</td>
            <td style="text-align: center; border: 1px solid #dddddd;"><a href="{link}" target="_blank">🔗</a></td>
            <td style="text-align: center; border: 1px solid #dddddd;">{source}</td>
        </tr>
        """
        counter += 1
    
    table_content += "</table>"
    
    # 이메일 푸터 수정 - 테이블과 같은 위치에 정렬
    footer_content = """
    <div style="width: 80%; max-width: 800px; min-width: 500px; margin: 30px auto 0; font-size: 14px; color: #666; border-top: 1px solid #eee; padding-top: 20px; text-align: left;">
        <p>-------------------------------------------------</p>
        <p>문의사항 또는 개선요청사항이 있다면, 정보기획팀 <a href="mailto:ckdpharmamorning@gmail.com">ckdpharmamorning@gmail.com</a> 으로 문의 주세요. (내선:332)</p>
        <p>뉴스레터 구독을 취소하시려면 <a href="{{ unsubscribe_link }}'>여기</a>를 클릭하세요.</p>
    </div>
    """
    
    # 최종 이메일 내용 생성
    email_content = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="font-family: 'Apple SD Gothic Neo', '맑은 고딕', 'Malgun Gothic', sans-serif; line-height: 1.6; color: #333; max-width: 800px; margin: 0 auto; padding: 20px;">
        {stats_header}
        {table_content}
        {footer_content}
    </body>
    </html>
    """
    
    return email_content

# 구독자 리스트를 배치로 나누는 유틸리티 함수
def batch_subscribers(subscribers, batch_size=5):
    """구독자 리스트를 지정된 크기의 배치로 나눕니다."""
    for i in range(0, len(subscribers), batch_size):
        yield subscribers[i:i + batch_size]

# SMTP 서버 연결 생성 함수
def create_smtp_connection():
    try:
        # SSL 컨텍스트 생성
        context = ssl.create_default_context()
        
        # SMTP 연결
        server = smtplib.SMTP('smtp.gmail.com', 587, timeout=30)
        
        # 전역 변수로 디버그 모드 제어
        if DEBUG_SMTP:
            server.set_debuglevel(1)
        else:
            server.set_debuglevel(0)
        
        # 소켓 연결 유지 설정
        server.sock.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
        
        # TCP keepalive 파라미터 설정 (리눅스)
        if hasattr(socket, 'TCP_KEEPIDLE') and hasattr(socket, 'TCP_KEEPINTVL') and hasattr(socket, 'TCP_KEEPCNT'):
            server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPIDLE, 60)
            server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPINTVL, 60)
            server.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPCNT, 5)
        
        server.ehlo()
        server.starttls(context=context)
        server.ehlo()
        server.login(EMAIL_ID, EMAIL_PW)
        return server
    except Exception as e:
        print(f"SMTP 연결 실패: {str(e)}")
        return None

# 뉴스 이메일 일괄 발송 함수
def send_headlines_email(all_headlines, subscribers, batch_size=3, delay_seconds=10):
    print("\n=== 헤드라인 이메일 발송 시작 ===")
    
    # 소켓 옵션 설정 (타임아웃 전에 연결 유지를 위한 keepalive 설정)
    socket.setdefaulttimeout(60)  # 전역 타임아웃 설정
    
    # 이메일 제목 생성 - 마지막 이모지를 일출로 변경
    today = datetime.now().strftime("%Y-%m-%d")
    subject = f"📰 오늘의 제약뉴스 모음 ({today}) 🌅"
    
    # 이메일 본문 생성
    email_content = generate_email_content(all_headlines)
    
    # 구독자별 이메일 발송
    success_count = 0
    fail_count = 0
    total_subscribers = len(subscribers)
    
    # 실패한 구독자 목록 추적
    failed_subscribers = []
    
    # 배치 처리를 위해 구독자 리스트 분할
    batches = list(batch_subscribers(subscribers, batch_size))
    total_batches = len(batches)
    
    print(f"총 {total_subscribers}명의 구독자를 {total_batches}개 배치로 처리합니다 (배치당 최대 {batch_size}명)")
    
    for batch_index, batch in enumerate(batches, 1):
        batch_start_time = datetime.now()
        batch_success = 0
        batch_fail = 0
        
        print(f"\n=== 배치 {batch_index}/{total_batches} 처리 시작 ({len(batch)}명) ===")
        
        # 각 배치마다 새로운 SMTP 연결 생성
        server = None
        max_retry = 3
        
        for retry in range(max_retry):
            server = create_smtp_connection()
            if server:
                print(f"SMTP 서버 연결 성공 (시도 {retry+1}/{max_retry})")
                break
            else:
                print(f"SMTP 서버 연결 실패 (시도 {retry+1}/{max_retry})")
                if retry < max_retry - 1:
                    time.sleep(5)
        
        if not server:
            print(f"배치 {batch_index} SMTP 연결 실패로 건너뜁니다")
            # 배치 전체를 실패 목록에 추가
            failed_subscribers.extend(batch)
            continue
        
        # 배치 내 각 구독자에게 이메일 전송
        for subscriber in batch:
            try:
                # 도메인 설정 (환경변수에서 가져오기)
                domain = os.environ.get('APP_DOMAIN', '10.4.31.151:80')
                
                # 구독 취소 링크 생성
                unsubscribe_url = f"http://{domain}/unsubscribe/{subscriber.unsubscribe_token}"
                
                # 멀티파트 메시지 생성
                msg = MIMEMultipart('mixed')
                msg['Subject'] = subject
                msg['From'] = EMAIL_SENDER
                msg['To'] = subscriber.email
                
                # 구독 취소 헤더 추가
                msg.add_header('List-Unsubscribe', f'<{unsubscribe_url}>')
                
                # 구독 취소 링크가 포함된 본문 생성
                personalized_content = email_content.replace("{{ unsubscribe_link }}", unsubscribe_url)
                personalized_content = personalized_content.replace("{ unsubscribe_link }", unsubscribe_url)
                
                # HTML 메시지 첨부
                html_part = MIMEText(personalized_content, 'html', 'utf-8')
                msg.attach(html_part)
                
                # 최대 3번 재시도 로직
                email_sent = False  # 이메일 발송 성공 여부 추적
                for attempt in range(3):
                    try:
                        server.send_message(msg)
                        print(f"{subscriber.email}로 이메일 전송 완료!")
                        success_count += 1
                        batch_success += 1
                        email_sent = True  # 성공 표시
                        time.sleep(1)  # 개별 이메일 사이에 지연 시간
                        break  # 성공하면 재시도 루프 탈출
                    except Exception as e:
                        if attempt == 2:  # 마지막 시도
                            print(f"{subscriber.email}로 이메일 전송 실패 (3회 시도 후): {str(e)}")
                            fail_count += 1
                        else:
                            print(f"{subscriber.email}로 이메일 전송 실패, 재시도 중 ({attempt+1}/3): {str(e)}")
                            # SMTP 연결 재설정 시도
                            try:
                                server.quit()
                                time.sleep(2 ** attempt)  # 지수 백오프
                                server = create_smtp_connection()
                                if not server:
                                    print("SMTP 재연결 실패, 다음 구독자로 넘어갑니다")
                                    fail_count += 1
                                    break
                            except Exception as conn_err:
                                print(f"SMTP 재연결 실패: {str(conn_err)}")
                                fail_count += 1
                                break
                
                # 3번 시도 후에도 실패했으면 실패 목록에 추가
                if not email_sent:
                    failed_subscribers.append(subscriber)
                    
            except Exception as e:
                print(f"구독자 {subscriber.email}에게 발송 준비 실패: {str(e)}")
                fail_count += 1
                failed_subscribers.append(subscriber)  # 실패 목록에 추가
        
        # SMTP 서버 연결 종료
        try:
            server.quit()
        except:
            pass
        
        batch_end_time = datetime.now()
        batch_duration = (batch_end_time - batch_start_time).total_seconds()
        
        # 배치 완료 로그
        print(f"=== 배치 {batch_index}/{total_batches} 완료: " +
              f"성공 {batch_success}건, 실패 {batch_fail}건, " +
              f"소요 시간 {batch_duration:.1f}초 ===")
        
        # 마지막 배치가 아니면 지연 시간 적용
        if batch_index < total_batches:
            print(f"다음 배치 전 {delay_seconds}초 대기 중...")
            time.sleep(delay_seconds)
            
    # 모든 배치 처리 완료 후 실패한 구독자에게 최종 재시도
    if failed_subscribers:
        print(f"\n=== 실패한 {len(failed_subscribers)}명에게 최종 재시도 시작 ===")
        
        # 실패한 구독자를 다시 배치로 나누어 처리
        retry_batches = list(batch_subscribers(failed_subscribers, batch_size))
        retry_total_batches = len(retry_batches)
        retry_success = 0
        retry_fail = 0
        
        for retry_batch_index, retry_batch in enumerate(retry_batches, 1):
            print(f"\n=== 최종 재시도 배치 {retry_batch_index}/{retry_total_batches} 처리 시작 ({len(retry_batch)}명) ===")
            
            # 새로운 SMTP 연결 생성
            retry_server = None
            for retry in range(max_retry):
                retry_server = create_smtp_connection()
                if retry_server:
                    print(f"최종 재시도 SMTP 서버 연결 성공 (시도 {retry+1}/{max_retry})")
                    break
                else:
                    print(f"최종 재시도 SMTP 서버 연결 실패 (시도 {retry+1}/{max_retry})")
                    if retry < max_retry - 1:
                        time.sleep(5)
            
            if not retry_server:
                print(f"최종 재시도 배치 {retry_batch_index} SMTP 연결 실패로 건너뜁니다")
                continue
                
            for subscriber in retry_batch:
                try:
                    # 도메인 설정
                    domain = os.environ.get('APP_DOMAIN', '10.4.31.151:80')
                    unsubscribe_url = f"http://{domain}/unsubscribe/{subscriber.unsubscribe_token}"
                    
                    # 메시지 생성
                    msg = MIMEMultipart('mixed')
                    msg['Subject'] = subject
                    msg['From'] = EMAIL_SENDER
                    msg['To'] = subscriber.email
                    msg.add_header('List-Unsubscribe', f'<{unsubscribe_url}>')
                    
                    # 본문 생성
                    personalized_content = email_content.replace("{{ unsubscribe_link }}", unsubscribe_url)
                    personalized_content = personalized_content.replace("{ unsubscribe_link }", unsubscribe_url)
                    html_part = MIMEText(personalized_content, 'html', 'utf-8')
                    msg.attach(html_part)
                    
                    # 최종 재시도 (3회)
                    final_sent = False
                    for attempt in range(3):
                        try:
                            retry_server.send_message(msg)
                            print(f"[최종 재시도] {subscriber.email}로 이메일 전송 완료!")
                            success_count += 1  # 전체 성공 카운트 증가
                            fail_count -= 1     # 이전에 실패로 카운트된 것 감소
                            retry_success += 1
                            final_sent = True
                            time.sleep(1)
                            break
                        except Exception as e:
                            if attempt == 2:  # 마지막 시도
                                print(f"[최종 재시도] {subscriber.email}로 이메일 전송 최종 실패: {str(e)}")
                                retry_fail += 1
                            else:
                                print(f"[최종 재시도] {subscriber.email}로 이메일 전송 실패, 재시도 중 ({attempt+1}/3): {str(e)}")
                                try:
                                    retry_server.quit()
                                    time.sleep(2 ** attempt)
                                    retry_server = create_smtp_connection()
                                    if not retry_server:
                                        print("[최종 재시도] SMTP 재연결 실패, 다음 구독자로 넘어갑니다")
                                        retry_fail += 1
                                        break
                                except Exception as conn_err:
                                    print(f"[최종 재시도] SMTP 재연결 실패: {str(conn_err)}")
                                    retry_fail += 1
                                    break
                except Exception as e:
                    print(f"[최종 재시도] 구독자 {subscriber.email}에게 발송 준비 실패: {str(e)}")
                    retry_fail += 1
            
            # SMTP 서버 종료
            try:
                retry_server.quit()
            except:
                pass
            
            # 마지막 배치가 아니면 지연 시간 적용
            if retry_batch_index < retry_total_batches:
                print(f"다음 최종 재시도 배치 전 {delay_seconds}초 대기 중...")
                time.sleep(delay_seconds)
        
        print(f"=== 최종 재시도 결과: 성공 {retry_success}건, 실패 {retry_fail}건 ===")
    
    print(f"\n=== 헤드라인 이메일 발송 완료: 성공 {success_count}건, 실패 {fail_count}건, 총 {total_batches}개 배치 ===")
    return {"success": success_count, "fail": fail_count, "batches": total_batches}

# 날짜 범위로 DB에서 뉴스 조회하는 함수
def get_headlines_by_date_range(db: Session, start_date, end_date):
    """날짜 범위 내의 뉴스 헤드라인을 DB에서 조회"""
    try:
        # datetime 객체로 변환
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
        
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            # 종료일은 해당일 23:59:59까지 포함
            end_date = end_date.replace(hour=23, minute=59, second=59)
        
        # DB 쿼리 (created_at 필드 사용)
        headlines = db.query(models.News).filter(
            models.News.created_at >= start_date,
            models.News.created_at <= end_date
        ).order_by(models.News.created_at.desc()).all()
        
        return headlines
    except Exception as e:
        print(f"뉴스 조회 중 오류 발생: {str(e)}")
        return []

# 주간 엑셀 보고서 생성 함수
def generate_weekly_excel_report(db: Session, start_date=None, end_date=None):
    """주간 뉴스 보고서 엑셀 파일 생성 - 각 날짜별로 별도 시트로 구성"""
    try:
        # 날짜 범위 설정 (기본: 이번 주 월요일~오늘)
        if not end_date:
            # 오늘 날짜
            today = datetime.now()
            end_date = today.replace(hour=23, minute=59, second=59)
        
        if not start_date:
            # 이번 주 월요일 계산 (오늘 - 요일값)
            today = datetime.now() if end_date is None else end_date
            days_since_monday = today.weekday()  # 월=0, 화=1, ..., 일=6
            this_monday = (today - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0)
            start_date = this_monday
        
        # 날짜별로 헤드라인 그룹화
        date_range = (end_date - start_date).days + 1

        # 날짜별 데이터 수집 - 날짜를 키로 하는 딕셔너리 생성
        daily_news = {}
        total_news_count = 0
        
        for day in range(date_range):
            current_date = start_date + timedelta(days=day)
            day_start = current_date.replace(hour=0, minute=0, second=0)
            day_end = current_date.replace(hour=23, minute=59, second=59)
            
            # 해당 날짜의 뉴스 가져오기
            headlines = get_headlines_by_date_range(db, day_start, day_end)
            
            if headlines:
                date_str = current_date.strftime('%Y-%m-%d')
                day_name = current_date.strftime('%A')  # 요일 이름 (영어)
                # 요일 이름 한글로 변환
                korean_day_names = {
                    'Monday': '월요일', 
                    'Tuesday': '화요일', 
                    'Wednesday': '수요일', 
                    'Thursday': '목요일',
                    'Friday': '금요일', 
                    'Saturday': '토요일', 
                    'Sunday': '일요일'
                }
                korean_day = korean_day_names.get(day_name, '')
                
                # 키를 '날짜 (요일)'로 설정
                sheet_name = f"{date_str} ({korean_day})"
                
                daily_news[sheet_name] = []
                
                for headline in headlines:
                    source_name = headline.source
                    # 출처 이름 한글로 변환
                    if source_name == "dailypharm":
                        source_name = "데일리팜"
                    elif source_name == "yakup":
                        source_name = "약업신문"
                    elif source_name == "hitnews":
                        source_name = "히트뉴스"
                    elif source_name == "kpanews":
                        source_name = "약사공론"
                    
                    daily_news[sheet_name].append({
                        "제목": headline.headline,
                        "출처": source_name,
                        "URL": headline.url
                    })
                    total_news_count += 1
        
        if total_news_count == 0:
            print("해당 기간에 저장된 뉴스가 없습니다.")
            return None
        
        # 보고서 폴더 생성
        os.makedirs("reports", exist_ok=True)
        
        # 파일명 설정
        report_date = datetime.now().strftime('%Y%m%d')
        period_text = f"{start_date.strftime('%Y-%m-%d')}~{end_date.strftime('%Y-%m-%d')}"
        filename = f"weekly_news_report_{report_date}.xlsx"
        filepath = os.path.join("reports", filename)
        
        # 엑셀 워크북 생성
        wb = Workbook()
        
        # 요약 시트 생성 (첫 시트)
        ws_summary = wb.active
        ws_summary.title = "주간 요약"
        
        # 요약 시트 제목
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"제약 뉴스 주간 리포트 ({period_text})"
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        # 요약 정보 추가
        ws_summary['A3'] = "날짜"
        ws_summary['B3'] = "뉴스 건수"
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'].font = Font(bold=True)
        
        # 요약 테이블 스타일
        for cell in [ws_summary['A3'], ws_summary['B3']]:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 각 날짜별 뉴스 건수 추가
        row = 4
        for sheet_name, news_items in daily_news.items():
            ws_summary[f'A{row}'] = sheet_name
            ws_summary[f'B{row}'] = len(news_items)
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='center')
            ws_summary[f'B{row}'].alignment = Alignment(horizontal='center')
            
            # 짝수 행 배경색
            if row % 2 == 0:
                ws_summary[f'A{row}'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                ws_summary[f'B{row}'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            
            row += 1
        
        # 요약 하단에 총계 추가
        ws_summary[f'A{row}'] = "총계"
        ws_summary[f'B{row}'] = total_news_count
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(bold=True)
        ws_summary[f'A{row}'].alignment = Alignment(horizontal='center')
        ws_summary[f'B{row}'].alignment = Alignment(horizontal='center')
        
        # 요약 시트 열 너비 조정
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        
        # 테두리 스타일 설정
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 요약 시트에 테두리 적용
        for r in range(3, row + 1):
            for c in range(1, 3):
                ws_summary.cell(row=r, column=c).border = thin_border
        
        # 각 날짜별로 시트 생성
        for sheet_name, news_items in daily_news.items():
            # 새 시트 생성
            ws = wb.create_sheet(title=sheet_name)
            
            # 제목 행 추가
            ws.merge_cells('A1:D1')
            ws['A1'] = f"제약 뉴스 모음 - {sheet_name}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # 열 제목 추가
            headers = ["번호", "제목", "출처", "바로가기"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(color="FFFFFF", bold=True)
            
            # 데이터 추가
            link_icon = "🔗"  # 링크 아이콘
            
            # 출처별로 정렬
            sorted_news = sorted(news_items, key=lambda x: x['출처'])
            
            for idx, news in enumerate(sorted_news, 1):
                row_idx = idx + 2  # 헤더 다음부터 시작
                
                # 번호, 제목, 출처 추가
                ws.cell(row=row_idx, column=1).value = idx
                ws.cell(row=row_idx, column=2).value = news['제목']
                ws.cell(row=row_idx, column=3).value = news['출처']
                
                # 링크 아이콘 추가
                link_cell = ws.cell(row=row_idx, column=4)
                link_cell.value = link_icon
                link_cell.hyperlink = news['URL']
                link_cell.font = Font(color="0563C1", bold=True)
                link_cell.alignment = Alignment(horizontal='center')
                
                # 짝수/홀수 행 배경색 설정
                if idx % 2 == 0:
                    for col in range(1, 5):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                        )
                
                # 테두리 추가
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).border = thin_border
            
            # 셀 정렬 설정
            for row in range(3, len(sorted_news) + 3):
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')  # 번호 - 중앙
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')    # 제목 - 왼쪽
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')  # 출처 - 중앙
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')  # 링크 - 중앙
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 8   # 번호
            ws.column_dimensions['B'].width = 70  # 제목
            ws.column_dimensions['C'].width = 12  # 출처
            ws.column_dimensions['D'].width = 10  # 바로가기 아이콘
            
            # 헤더 행에 테두리 적용
            for col in range(1, 5):
                ws.cell(row=2, column=col).border = thin_border
        
        # 파일 저장
        wb.save(filepath)
        print(f"엑셀 보고서가 생성되었습니다: {filepath}")
        
        # 파일 생성 확인
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"생성된 파일 크기: {file_size} 바이트")
            
            # 파일 확장자 확인
            _, file_ext = os.path.splitext(filepath)
            if file_ext.lower() != '.xlsx':
                print(f"경고: 파일 확장자가 .xlsx가 아닙니다. 현재 확장자: {file_ext}")
                # 올바른 확장자로 파일명 변경
                new_filepath = f"{os.path.splitext(filepath)[0]}.xlsx"
                os.rename(filepath, new_filepath)
                filepath = new_filepath
                print(f"파일명을 {filepath}로 변경했습니다.")
                
            return filepath
        else:
            print(f"오류: 파일 {filepath}가 생성되지 않았습니다.")
            return None
            
    except Exception as e:
        print(f"엑셀 보고서 생성 중 오류 발생: {str(e)}")
        print(traceback.format_exc())
        return None

# 주간 리포트 이메일 발송 함수
def send_weekly_report_email(db: Session, recipients=None, start_date=None, end_date=None, batch_size=5, delay_seconds=3):
    """주간 뉴스 리포트 엑셀 파일을 생성하고 이메일로 발송"""
    try:
        # 엑셀 파일 생성
        excel_file = generate_weekly_excel_report(db, start_date, end_date)
        if not excel_file:
            return {
                "success": False, 
                "message": "엑셀 리포트 생성 실패. 해당 기간에 뉴스가 없을 수 있습니다."
            }
        
        # 기간 텍스트 생성
        if start_date and end_date:
            period_text = f"{start_date.strftime('%Y.%m.%d')}~{end_date.strftime('%Y.%m.%d')}"
        else:
            # 날짜 범위 설정 (기본: 이번 주 월요일~오늘)
            today = datetime.now()
            days_since_monday = today.weekday()
            this_monday = (today - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0)
            period_text = f"{this_monday.strftime('%Y.%m.%d')}~{today.strftime('%Y.%m.%d')}"
        
        # 이메일 제목 설정 (이모지 추가)
        subject = f"📊 주간 제약뉴스 리포트 ({datetime.now().strftime('%Y-%m-%d')}) 📈"
        
        # 이메일 본문 생성 (간결한 디자인)
        email_content = f"""
        <div style='font-family: Arial, sans-serif;'>
            <h2>📊 제약뉴스 주간 리포트</h2>
            
            <div style='font-size: 1.0em; color: #333; margin: 20px 0;'>
                📅 수집 기간: {period_text} 기간의 뉴스
            </div>
            
            <p>첨부된 엑셀 파일을 확인해 주세요. 💼</p>
            <p>즐거운 하루 되세요! ✨</p>
            <br>
            <small>문의사항 또는 개선요청사항이 있다면, 정보기획팀 <a href='mailto:ckdpharmamorning@gmail.com'>ckdpharmamorning@gmail.com</a> 으로 문의 주세요. (내선:332)</small>
            <br>
            <small>뉴스레터 구독을 취소하시려면 <a href='{{ unsubscribe_link }}'>여기</a>를 클릭하세요.</small>
        </div>
        """
        
        # 구독자 목록 가져오기 (recipients가 None인 경우)
        if recipients is None:
            recipients = db.query(models.Subscriber).filter(models.Subscriber.is_active == True).all()
            
        if not recipients:
            return {
                "success": False,
                "message": "활성화된 구독자가 없습니다."
            }
        
        # 첨부 파일명 설정
        attachment_filename = f"제약_뉴스_주간리포트_{period_text.replace('.', '').replace('~', '-')}.xlsx"
        
        # 이메일 발송 (배치 처리)
        success_count = 0
        fail_count = 0
        total_subscribers = len(recipients)
        
        # 실패한 구독자 추적을 위한 리스트 추가
        failed_recipients = []
        
        # 배치 처리를 위해 구독자 리스트 분할
        batches = list(batch_subscribers(recipients, batch_size))
        total_batches = len(batches)
        
        print(f"주간 리포트: 총 {total_subscribers}명의 구독자를 {total_batches}개 배치로 처리합니다 (배치당 최대 {batch_size}명)")
        
        for batch_index, batch in enumerate(batches, 1):
            batch_start_time = datetime.now()
            batch_success = 0
            batch_fail = 0
            
            print(f"\n=== 주간 리포트 배치 {batch_index}/{total_batches} 처리 시작 ({len(batch)}명) ===")
            
            for recipient in batch:
                try:
                    # 이메일 발송
                    result = send_email(
                        subscriber=recipient, 
                        subject=subject, 
                        content=email_content,
                        attachment_path=excel_file,
                        attachment_filename=attachment_filename
                    )
                    
                    if result:
                        success_count += 1
                        batch_success += 1
                    else:
                        fail_count += 1
                        batch_fail += 1
                        # 실패한 구독자 리스트에 추가
                        failed_recipients.append(recipient)
                except Exception as e:
                    print(f"구독자 {recipient.email} 처리 중 오류: {str(e)}")
                    fail_count += 1
                    batch_fail += 1
                    # 실패한 구독자 리스트에 추가
                    failed_recipients.append(recipient)
            
            batch_end_time = datetime.now()
            batch_duration = (batch_end_time - batch_start_time).total_seconds()
            
            # 배치 완료 로그
            print(f"=== 주간 리포트 배치 {batch_index}/{total_batches} 완료: " +
                  f"성공 {batch_success}건, 실패 {batch_fail}건, " +
                  f"소요 시간 {batch_duration:.1f}초 ===")
            
            # 마지막 배치가 아니면 지연 시간 적용
            if batch_index < total_batches:
                print(f"다음 배치 전 {delay_seconds}초 대기 중...")
                time.sleep(delay_seconds)
        
        # 실패한 구독자 재시도 로직 추가
        if failed_recipients:
            print(f"\n=== 주간 리포트: 실패한 {len(failed_recipients)}명에게 최종 재시도 시작 ===")
            
            # 실패한 구독자를 다시 배치로 나누어 처리
            retry_batches = list(batch_subscribers(failed_recipients, batch_size))
            retry_total_batches = len(retry_batches)
            retry_success = 0
            retry_fail = 0
            
            for retry_batch_index, retry_batch in enumerate(retry_batches, 1):
                print(f"\n=== 최종 재시도 배치 {retry_batch_index}/{retry_total_batches} 처리 시작 ({len(retry_batch)}명) ===")
                
                for recipient in retry_batch:
                    try:
                        # 이메일 재발송
                        result = send_email(
                            subscriber=recipient, 
                            subject=subject, 
                            content=email_content,
                            attachment_path=excel_file,
                            attachment_filename=attachment_filename
                        )
                        
                        if result:
                            print(f"[주간 리포트 최종 재시도] {recipient.email}로 이메일 전송 완료!")
                            success_count += 1  # 전체 성공 카운트 증가
                            fail_count -= 1     # 이전에 실패로 카운트된 것 감소
                            retry_success += 1
                        else:
                            print(f"[주간 리포트 최종 재시도] {recipient.email}로 이메일 전송 최종 실패")
                            retry_fail += 1
                    except Exception as e:
                        print(f"[주간 리포트 최종 재시도] 구독자 {recipient.email} 재처리 중 오류: {str(e)}")
                        retry_fail += 1
                
                # 마지막 배치가 아니면 지연 시간 적용
                if retry_batch_index < retry_total_batches:
                    print(f"다음 최종 재시도 배치 전 {delay_seconds}초 대기 중...")
                    time.sleep(delay_seconds)
            
            print(f"=== 주간 리포트 최종 재시도 결과: 성공 {retry_success}건, 실패 {retry_fail}건 ===")
        
        return {
            "success": True,
            "message": f"주간 리포트 이메일 발송 완료: 성공 {success_count}건, 실패 {fail_count}건, 총 {total_batches}개 배치",
            "file": excel_file,
            "success_count": success_count,
            "fail_count": fail_count,
            "batches": total_batches
        }
        
    except Exception as e:
        print(f"주간 리포트 이메일 발송 중 오류 발생: {str(e)}")
        print(traceback.format_exc())
        return {"success": False, "message": f"오류 발생: {str(e)}"}


